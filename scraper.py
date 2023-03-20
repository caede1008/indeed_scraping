import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.by import By
import openpyxl
from tkinter import *
from tkinter import ttk
import requests

# GUI設定
root = Tk()
root.title('Indeedスクレイピングツール')
root.geometry("400x300+100+300")
frame = ttk.Frame(root, padding=16)

keywordlbl = ttk.Label(text='キーワード')
keywordlbl.place(x=20, y=30)

txtbox = ttk.Entry()
txtbox.configure(state='normal', width=35)
txtbox.place(x=110, y=30)

wkplacelbl = ttk.Label(text='勤務地')
wkplacelbl.place(x=20, y=60)

wkplacebox = ttk.Entry()
wkplacebox.configure(state='normal', width=35)
wkplacebox.place(x=110, y=60)

maxlbl = ttk.Label(text='上限件数')
maxlbl.place(x=20, y=90)

maxbox = ttk.Entry()
maxbox.configure(state='normal', width=10)
maxbox.insert(0, "10")
maxbox.place(x=110, y=90)

scrapingbutton = ttk.Button(root, text="スクレイピング実行", command=lambda:main(), width=40)
scrapingbutton.pack(pady=130)

def main():
    BrowserPath=ResourcePath("./browser/chrome.exe") # ブラウザ
    DriverPath=ResourcePath("./driver/chromedriver.exe") # ウェブドライバ

    # ウェブドライバ設定
    options=Options()
    options.binary_location=BrowserPath
    # options.add_argument("--headless") # 動きを見たい場合はコメントアウトする。
    driver=webdriver.Chrome(DriverPath, options=options)

    # エクセル準備
    path = './indeed_list.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb["list"]
    exs = wb["exlist"]
    rownumber = 2

    # スクレイピング準備
    url = 'https://jp.indeed.com/?from=gnav-homepage'
    driver.get(url)
    time.sleep(2)
    ttlcnt = 0

    # キーワード設定
    keywordtxt = txtbox.get()
    if keywordtxt.__contains__(','):
        wklist = keywordtxt.split(',')
        keywordtxt = ''
        for keyword in wklist:
            keywordtxt = keywordtxt + ' ' + keyword
    if len(keywordtxt) != 0:
        kwsearchbox = driver.find_element(By.XPATH, "//*[@id='text-input-what']")
        kwsearchbox.send_keys(keywordtxt)

    # 勤務地
    wkplacetext = wkplacebox.get()
    if wkplacetext.__contains__(','):
        wklist = wkplacetext.split(',')
        wkplacetext = ''
        for wktxt in wklist:
            wkplacetext = wkplacetext + ' ' + wktxt
    if len(wkplacetext) != 0:
        wkp_searchbox = driver.find_element(By.XPATH, "//*[@id='text-input-where']")
        wkp_searchbox.send_keys(wkplacetext)

    # 除外ワード
    exstr = ''
    firststr = exs.cell(1, 1).value
    if firststr != None:
        maxrow = exs.max_row + 1
        for rn in range(1, maxrow):
            exword = exs.cell(rn, 1).value
            exstr = exstr + " -" + exword
        if len(exstr) != 0:
            kwsearchbox = driver.find_element(By.XPATH, "//*[@id='text-input-what']")
            kwsearchbox.send_keys(exstr)

    # 件数上限設定
    maxcnt = int(maxbox.get()) - 1

    # 検索ボタン押下
    searchbutton = driver.find_element(By.XPATH, "//*[@id='jobsearch']/button")
    searchbutton.click()
    time.sleep(3)

    # 掲載日設定
    wkdatebutton = []
    wkdatebutton = driver.find_elements(By.CLASS_NAME, "yosegi-FilterPill-pill")
    databutton = wkdatebutton[0]
    databutton.click()
    time.sleep(1)
    wktwoweekbutton = []
    wktwoweekbutton = driver.find_elements(By.CLASS_NAME, "yosegi-FilterPill-dropdownListItemLink")
    twoweekbutton = wktwoweekbutton[3]
    twoweekbutton.click()
    time.sleep(2)

    # 掲載順設定
    wkorderbutton = driver.find_elements(By.LINK_TEXT, "日付順")
    if len(wkorderbutton) != 0:
        orderbutton = wkorderbutton[0]
        orderbutton.click()
        time.sleep(3)

    while True:

        # 変数リセット
        companynames = []
        contents = []

        # Cookie同意
        ckbutton = driver.find_elements(By.CLASS_NAME, "gnav-CookiePrivacyNoticeButton")
        if len(ckbutton) != 0:
            ckbutton[0].click()

        # 各案件ボタン
        buttons = []
        buttons = driver.find_elements(By.CLASS_NAME, "css-1m4cuuf.e37uo190")

        # 掲載時間/日取得
        datehours = []
        wkdatehours = driver.find_elements(By.CLASS_NAME, "date")
        for wkdatehour in wkdatehours:
            if len(wkdatehour.text) != 0:
                datehours.append(wkdatehour.text[7:])

        # スクレイピング実行
        spcnt = 0
        wkcompanynames = driver.find_elements(By.CLASS_NAME, "companyName")
        for wkcompanyname in wkcompanynames:
            companynames.append(wkcompanyname.text)
            spcnt += 1
            ttlcnt += 1
            if ttlcnt > maxcnt:
                break

        for i in range(0, spcnt):
            buttons[i].click()
            time.sleep(2)
            wkcontents = driver.find_elements(By.CLASS_NAME, "jobsearch-JobComponent-embeddedBody")
            if len(wkcontents) != 0:
                contents.append(wkcontents[0].text)
            else:
                contents.append('無し')

        # Excel入力
        idx = 0
        for companyname in companynames:
            ws.cell(rownumber, 1).value = companyname
            ws.cell(rownumber, 2).value = datehours[idx]
            ws.cell(rownumber, 3).value = contents[idx]
            idx += 1
            rownumber += 1

        wb.save(path)
        wb.close()

        # 500件チェック
        if ttlcnt > maxcnt:
            break

        # 次ページ移行
        nextbutton = driver.find_elements(By.CLASS_NAME, "css-13p07ha.e8ju0x50")
        if len(nextbutton)>1:
            nextbutton[1].click()
        else:
            nextbutton[0].click()
        time.sleep(5)

    # クローズ処理
    time.sleep(5)
    driver.close()
    driver.quit()

def ResourcePath(relativePath):
    try:
        basePath=sys._MEIPASS
    except Exception:
        basePath=os.path.dirname(__file__)
    return os.path.join(basePath, relativePath)

#if __name__=="__main__":
    #main()


root.mainloop()
