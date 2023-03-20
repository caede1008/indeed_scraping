import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.by import By
import openpyxl
from tkinter import *
from tkinter import ttk
import requests

# エクセル準備
path = './indeed_list.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb["list"]
exs2 = wb["exlist2"]
rownumber = 2

# 除外キーワード削除処理
dellist = []
dellistfl = []
delchk = False
listrow = 2
delmax = exs2.max_row
listmaxrow = ws.max_row
wkdelrows = ""
delrows = ""

for listrow in range(2, listmaxrow + 1):
    delchk = False
    liststr = ws.cell(listrow, 1).value
    if liststr != None:
        # 除外シート設定
        for delrow in range(1, delmax + 1):
            delstr = exs2.cell(delrow, 1).value
            if delstr != None:
                if liststr.__contains__(delstr):
                    delchk = True
                    break
        if delchk == True:
            dellist.append(listrow)

# 行削除
delcnt = 0
if len(dellist) !=0:
    dellist.sort()
    for delrownum in dellist:
        ws.delete_rows(delrownum - delcnt)
        delcnt += 1

wb.save(path)
wb.close()
